"""
Excel 寫入模組。

工作表：
  單打紀錄      ← 所有單打對局
  雙打紀錄      ← 所有雙打對局
  強力排組（單打）← 單打第一名
  強力排組（雙打）← 雙打第一名

欄位順序：
  日期時間 | 版本 | 模式 | 英雄 | 技能1 | 技能2 | 飾品1 | 飾品2 |
  名次 | 最終排組（隨從+數值）| 回合數 | 最高金幣 | 對局時長 | 對手英雄 | 備註
"""
import os
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(__file__)
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
EXCEL_PATH  = os.path.join(OUTPUT_DIR, "hs_bg_records.xlsx")

HEADERS = [
    "日期時間", "版本", "模式", "英雄", "隊友英雄",
    "技能1", "技能2", "飾品1", "飾品2",
    "名次", "最終排組（隨從）",
    "回合數", "最高金幣", "對局時長", "對手英雄", "對手排組", "備註",
]
COL_WIDTHS = [18, 12, 8, 22, 22, 20, 20, 22, 22, 6, 80, 8, 8, 10, 50, 80, 20]

SHEET_SOLO      = "單打紀錄"
SHEET_DUO       = "雙打紀錄"
SHEET_SOLO_TOP  = "強力排組（單打）"
SHEET_DUO_TOP   = "強力排組（雙打）"
ALL_SHEETS      = [SHEET_SOLO, SHEET_DUO, SHEET_SOLO_TOP, SHEET_DUO_TOP]

# 名次顏色
PLACE_COLORS = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _wb() -> openpyxl.Workbook:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        return openpyxl.load_workbook(EXCEL_PATH)
    wb = openpyxl.Workbook()
    # 刪除預設空白 sheet
    default = wb.active
    wb.remove(default)
    for name in ALL_SHEETS:
        _setup_sheet(wb.create_sheet(name))
    return wb


def _setup_sheet(ws):
    ws.append(HEADERS)
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=1, column=i)
        c.font   = HEADER_FONT
        c.fill   = HEADER_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _row(
    start_time: datetime,
    build_version: str,
    game_mode: str,
    hero_name: str,
    teammate_hero_name: str,
    power_names: list[str],
    trinket_names: list[str],
    placement: int,
    board_strs: list[str],
    turn_count: int,
    max_gold: int,
    duration_seconds: int,
    opponent_names: list[str],
    opponent_board_strs: list[str],
    note: str = "",
) -> list:
    mode_label = "雙打" if game_mode == "duo" else "單打"
    p1 = power_names[0] if len(power_names) > 0 else ""
    p2 = power_names[1] if len(power_names) > 1 else ""
    t1 = trinket_names[0] if len(trinket_names) > 0 else ""
    t2 = trinket_names[1] if len(trinket_names) > 1 else ""
    board_text = "、".join(board_strs) if board_strs else "（無資料）"
    mins, secs = divmod(duration_seconds, 60)
    duration_str = f"{mins}分{secs:02d}秒"
    opponents = "、".join(opponent_names) if opponent_names else ""
    opp_boards = "\n".join(opponent_board_strs) if opponent_board_strs else ""
    return [
        start_time.strftime("%Y-%m-%d %H:%M"),
        build_version, mode_label, hero_name, teammate_hero_name,
        p1, p2, t1, t2,
        placement, board_text,
        turn_count, max_gold, duration_str,
        opponents, opp_boards, note,
    ]


def _style_row(ws, row_idx: int, placement: int):
    color = PLACE_COLORS.get(placement, "FFFFFF")
    fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
    wrap_cols = {
        HEADERS.index("最終排組（隨從）") + 1,
        HEADERS.index("對手英雄") + 1,
        HEADERS.index("對手排組") + 1,
    }
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row_idx, column=col)
        c.fill      = fill
        c.border    = THIN
        c.alignment = Alignment(vertical="center", wrap_text=(col in wrap_cols))
    ws.row_dimensions[row_idx].height = 18


def append_record(
    start_time: datetime,
    build_version: str,
    game_mode: str,
    hero_name: str,
    power_names: list[str],
    trinket_names: list[str],
    placement: int,
    board_strs: list[str],
    turn_count: int,
    max_gold: int,
    duration_seconds: int,
    opponent_names: list[str],
    opponent_board_strs: list[str],
    teammate_hero_name: str = "",
    note: str = "",
) -> bool:
    """將對局寫入 Excel。第一名同時寫入對應強力排組分頁。回傳是否成功。"""
    try:
        wb = _wb()
        row_data = _row(
            start_time, build_version, game_mode, hero_name, teammate_hero_name,
            power_names, trinket_names, placement, board_strs,
            turn_count, max_gold, duration_seconds,
            opponent_names, opponent_board_strs, note,
        )

        # 決定寫入的主分頁
        main_sheet = SHEET_DUO if game_mode == "duo" else SHEET_SOLO
        ws = wb[main_sheet]
        ws.append(row_data)
        _style_row(ws, ws.max_row, placement)

        # 第一名 → 強力排組
        if placement == 1:
            top_sheet = SHEET_DUO_TOP if game_mode == "duo" else SHEET_SOLO_TOP
            ws_top = wb[top_sheet]
            ws_top.append(row_data)
            _style_row(ws_top, ws_top.max_row, placement)
            print(f"[excel] OK 加入強力排組（{top_sheet}）：{hero_name}")

        wb.save(EXCEL_PATH)
        return True

    except PermissionError:
        print("[excel] FAIL Excel 已開啟，請關閉後重試")
        return False
    except Exception as e:
        print(f"[excel] FAIL 錯誤：{e}")
        return False


def get_record_counts() -> dict:
    if not os.path.exists(EXCEL_PATH):
        return {s: 0 for s in ALL_SHEETS}
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    result = {}
    for name in ALL_SHEETS:
        ws = wb[name]
        result[name] = max(0, ws.max_row - 1)
    wb.close()
    return result

