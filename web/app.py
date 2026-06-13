"""
Flask 後端：提供對局紀錄 API 與網頁介面。
"""
import sys, os
import re
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import json
import sqlite3
from flask import Flask, render_template, jsonify, request, redirect, url_for
import openpyxl
from werkzeug.utils import secure_filename

BG_TRINKETS_CACHE = os.path.join(os.path.dirname(__file__), "..", "data", "bg_trinkets_cache.json")
BG_MINIONS_CACHE  = os.path.join(os.path.dirname(__file__), "..", "data", "bg_minions_cache.json")
BG_SPELLS_CACHE   = os.path.join(os.path.dirname(__file__), "..", "data", "bg_spells_cache.json")
BG_CONFIG_PATH    = os.path.join(os.path.dirname(__file__), "..", "data", "bg_config.json")
BG_HEROES_CACHE   = os.path.join(os.path.dirname(__file__), "..", "data", "bg_heroes_cache.json")
BG_COMPS_CACHE    = os.path.join(os.path.dirname(__file__), "..", "data", "bg_comps.json")
CARDS_CACHE       = os.path.join(os.path.dirname(__file__), "..", "data", "cards_cache.json")
HS_CARDS_FULL     = os.path.join(os.path.dirname(__file__), "..", "data", "hs_cards_full.json")
HS_BG_HEROES      = os.path.join(os.path.dirname(__file__), "..", "data", "hs_bg_heroes.json")
DB_PATH          = os.path.join(os.path.dirname(__file__), "..", "data", "records.db")

UPLOAD_DIR  = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB

# 公開模式：設定環境變數 PUBLIC_MODE=1 隱藏私人紀錄頁面
PUBLIC_MODE = os.environ.get("PUBLIC_MODE", "0") == "1"
# 管理員 token：設定環境變數 ADMIN_TOKEN=your_secret
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "")


@app.context_processor
def inject_globals():
    return {"public_mode": PUBLIC_MODE}


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_db():
    """確保資料表存在（雲端首次啟動時建立空資料表）。"""
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS games (
                game_id TEXT PRIMARY KEY,
                start_time TEXT NOT NULL,
                end_time TEXT,
                build_version TEXT DEFAULT '',
                game_mode TEXT DEFAULT 'solo',
                hero_card_id TEXT DEFAULT '',
                hero_name TEXT DEFAULT '',
                hero_power_ids TEXT DEFAULT '[]',
                hero_power_names TEXT DEFAULT '[]',
                trinket_ids TEXT DEFAULT '[]',
                trinket_names TEXT DEFAULT '[]',
                placement INTEGER DEFAULT 0,
                final_board TEXT DEFAULT '[]',
                turn_count INTEGER DEFAULT 0,
                max_gold INTEGER DEFAULT 0,
                duration_seconds INTEGER DEFAULT 0,
                opponent_heroes TEXT DEFAULT '[]',
                opponent_boards TEXT DEFAULT '{}',
                exported INTEGER DEFAULT 0,
                teammate_hero_card_id TEXT DEFAULT '',
                teammate_hero_name TEXT DEFAULT '',
                penultimate_board TEXT DEFAULT '[]'
            )
        """)
        conn.commit()


_ensure_db()

@app.route("/")
def index():
    if PUBLIC_MODE:
        return redirect(url_for("tier_list_page"))
    return render_template("index.html")


@app.route("/api/games")
def api_games():
    mode   = request.args.get("mode", "all")   # all / solo / duo
    place  = request.args.get("place", "all")  # all / 1 / 1-4
    sort   = request.args.get("sort", "date_desc")
    page   = int(request.args.get("page", 1))
    limit  = int(request.args.get("limit", 20))

    conditions = []
    params = []

    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)

    if place == "1":
        conditions.append("placement = 1")
    elif place == "top4":
        conditions.append("placement <= 4")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    order_map = {
        "date_desc":  "start_time DESC",
        "date_asc":   "start_time ASC",
        "place_asc":  "placement ASC, start_time DESC",
        "place_desc": "placement DESC, start_time DESC",
    }
    order = order_map.get(sort, "start_time DESC")

    offset = (page - 1) * limit

    with get_db() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM games {where}", params).fetchone()[0]
        rows  = conn.execute(
            f"SELECT * FROM games {where} ORDER BY {order} LIMIT ? OFFSET ?",
            params + [limit, offset]
        ).fetchall()

    games = []
    for r in rows:
        g = dict(r)
        for key in ("hero_power_ids","hero_power_names","trinket_ids","trinket_names",
                    "final_board","penultimate_board","opponent_heroes"):
            try:
                g[key] = json.loads(g.get(key) or "[]")
            except:
                g[key] = []
        try:
            g["opponent_boards"] = json.loads(g.get("opponent_boards") or "{}")
        except:
            g["opponent_boards"] = {}
        # 計算時長文字
        s = g.get("duration_seconds", 0) or 0
        g["duration_str"] = f"{s//60}分{s%60:02d}秒"
        games.append(g)

    return jsonify({"games": games, "total": total, "page": page, "limit": limit})


@app.route("/api/stats")
def api_stats():
    with get_db() as conn:
        total  = conn.execute("SELECT COUNT(*) FROM games").fetchone()[0]
        solo   = conn.execute("SELECT COUNT(*) FROM games WHERE game_mode='solo'").fetchone()[0]
        duo    = conn.execute("SELECT COUNT(*) FROM games WHERE game_mode='duo'").fetchone()[0]
        first  = conn.execute("SELECT COUNT(*) FROM games WHERE placement=1").fetchone()[0]
        avg_q  = conn.execute("SELECT AVG(placement) FROM games").fetchone()[0]
        avg_place = round(avg_q, 2) if avg_q else 0

        # 最常使用英雄（前5）
        hero_rows = conn.execute(
            "SELECT hero_name, COUNT(*) as cnt, AVG(placement) as avg_p "
            "FROM games WHERE hero_name != '' GROUP BY hero_name ORDER BY cnt DESC LIMIT 5"
        ).fetchall()
        top_heroes = [{"name": r["hero_name"], "games": r["cnt"],
                       "avg": round(r["avg_p"], 2)} for r in hero_rows]

    return jsonify({
        "total": total, "solo": solo, "duo": duo,
        "first": first, "avg_place": avg_place,
        "top_heroes": top_heroes
    })


@app.route("/top-builds")
def top_builds():
    return render_template("top_builds.html")


@app.route("/comps")
def comps():
    return render_template("comps.html")


@app.route("/api/comps")
def api_comps():
    import re as _re
    mode = request.args.get("mode", "all")

    conditions = ["placement > 0"]
    params: list = []
    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)
    where = "WHERE " + " AND ".join(conditions)

    with get_db() as conn:
        rows = conn.execute(
            f"SELECT * FROM games {where} ORDER BY placement ASC, start_time DESC",
            params
        ).fetchall()

    boards = []
    minion_map: dict = {}
    hero_map: dict = {}

    for row in rows:
        g = dict(row)
        placement = g.get("placement", 0)
        if not placement:
            continue

        try:
            board = json.loads(g.get("final_board") or "[]")
        except Exception:
            board = []

        s_val = g.get("duration_seconds", 0) or 0
        boards.append({
            "game_id":     g["game_id"],
            "hero_name":   g["hero_name"],
            "hero_card_id": g["hero_card_id"] or "",
            "placement":   placement,
            "game_mode":   g["game_mode"],
            "start_time":  g["start_time"],
            "turn_count":  g["turn_count"],
            "max_gold":    g["max_gold"],
            "duration_str": f"{s_val//60}分{s_val%60:02d}秒",
            "board":       board,
        })

        # per-minion stats
        for m in board:
            cid = m.get("card_id", "")
            if not cid:
                continue
            base = _re.sub(r"_G$", "", cid)
            name = m.get("name", base)
            if base not in minion_map:
                minion_map[base] = {"card_id": base, "name": name,
                                    "count": 0, "placements": [],
                                    "golden_count": 0, "top4": 0}
            minion_map[base]["count"] += 1
            minion_map[base]["placements"].append(placement)
            if m.get("golden") or cid.endswith("_G"):
                minion_map[base]["golden_count"] += 1
            if placement <= 4:
                minion_map[base]["top4"] += 1

        # per-hero stats
        hname = g.get("hero_name", "")
        hcid  = g.get("hero_card_id", "")
        if hname:
            if hname not in hero_map:
                hero_map[hname] = {"name": hname, "card_id": hcid,
                                   "count": 0, "placements": [],
                                   "wins": 0, "top4": 0}
            hero_map[hname]["count"] += 1
            hero_map[hname]["placements"].append(placement)
            if placement == 1: hero_map[hname]["wins"] += 1
            if placement <= 4: hero_map[hname]["top4"] += 1

    top_minions = []
    for s in minion_map.values():
        c = s["count"]
        avg_p = sum(s["placements"]) / c
        top_minions.append({
            "card_id":      s["card_id"],
            "name":         s["name"],
            "count":        c,
            "avg_placement": round(avg_p, 2),
            "golden_count": s["golden_count"],
            "top4":         s["top4"],
            "top4_rate":    round(s["top4"] / c * 100, 1),
        })
    top_minions.sort(key=lambda x: (x["avg_placement"], -x["count"]))

    heroes = []
    for s in hero_map.values():
        c = s["count"]
        avg_p = sum(s["placements"]) / c
        heroes.append({
            "name":         s["name"],
            "card_id":      s["card_id"],
            "count":        c,
            "avg_placement": round(avg_p, 2),
            "wins":         s["wins"],
            "top4":         s["top4"],
            "win_rate":     round(s["wins"] / c * 100, 1),
            "top4_rate":    round(s["top4"] / c * 100, 1),
        })
    heroes.sort(key=lambda x: x["avg_placement"])

    return jsonify({
        "boards":      boards,
        "top_minions": top_minions,
        "heroes":      heroes,
        "total":       len(boards),
    })


@app.route("/trinkets")
def trinkets_page():
    return render_template("trinkets.html")


@app.route("/api/trinkets")
def api_trinkets():
    tier_filter = request.args.get("tier", "all")   # all / lesser / greater
    mode = request.args.get("mode", "all")

    # ── Load trinket DB ──
    if os.path.exists(BG_TRINKETS_CACHE):
        with open(BG_TRINKETS_CACHE, encoding="utf-8") as f:
            trinket_db = {t["id"]: t for t in json.load(f)}
    else:
        trinket_db = {}

    # ── Personal stats ──
    conditions = ["placement > 0"]
    params: list = []
    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)
    where = "WHERE " + " AND ".join(conditions)

    personal: dict = {}
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT trinket_ids, placement FROM games {where}", params
        ).fetchall()

    for row in rows:
        placement = row["placement"]
        try:
            ids = json.loads(row["trinket_ids"] or "[]")
        except Exception:
            ids = []
        for tid in ids:
            if not tid:
                continue
            if tid not in personal:
                personal[tid] = {"count": 0, "placements": [], "top4": 0, "best": 99}
            e = personal[tid]
            e["count"] += 1
            e["placements"].append(placement)
            if placement <= 4: e["top4"] += 1
            if placement < e["best"]: e["best"] = placement

    def _stats(tid):
        p = personal.get(tid, {})
        c = p.get("count", 0)
        return {
            "seen":          c,
            "avg_placement": round(sum(p["placements"]) / c, 2) if c else None,
            "best":          p.get("best") if c else None,
            "top4":          p.get("top4", 0),
            "top4_rate":     round(p["top4"] / c * 100, 1) if c else None,
        }

    result = []
    for tid, t in trinket_db.items():
        if tier_filter != "all" and t["tier"] != tier_filter:
            continue
        s = _stats(tid)
        # also check variant stats
        vs = _stats(t["variant_id"]) if t.get("variant_id") else {}
        result.append({
            "id":           tid,
            "name":         t["name"],
            "tier":         t["tier"],
            "text":         t["text"],
            "variant_id":   t.get("variant_id"),
            "variant_name": t.get("variant_name", ""),
            "variant_text": t.get("variant_text", ""),
            **s,
            "variant_seen":          vs.get("seen", 0),
            "variant_avg_placement": vs.get("avg_placement"),
            "variant_best":          vs.get("best"),
            "variant_top4_rate":     vs.get("top4_rate"),
        })

    result.sort(key=lambda x: (x["tier"], x["name"]))
    return jsonify({"trinkets": result, "total": len(result)})


@app.route("/minions")
def minions_page():
    return render_template("minions.html")


@app.route("/api/minions")
def api_minions():

    # ── Load full BG minion DB ──
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            raw_list = json.load(f)
    else:
        raw_list = []

    # ── Merge golden (_G / _Gt / _Gt2) into base card ──
    # Build base dict first, then attach golden data
    bg_db = {}
    golden_db = {}
    for m in raw_list:
        cid = m["id"]
        # Detect golden suffix: _G, _Gt, _Gt2 etc.
        if cid.endswith("_G"):
            golden_db[cid[:-2]] = m          # BG28_503_G → BG28_503
        elif re.search(r"_Gt\d*$", cid):
            # Token golden: BG29_875_Gt → BG29_875t, BG29_875_Gt2 → BG29_875t2
            base = re.sub(r"_G(t\d*)$", r"\1", cid)
            golden_db[base] = m
        else:
            bg_db[cid] = m

    # Attach golden stats to base cards
    for base_id, gcard in golden_db.items():
        if base_id in bg_db:
            bg_db[base_id]["golden_attack"] = gcard.get("attack", 0)
            bg_db[base_id]["golden_health"] = gcard.get("health", 0)
            bg_db[base_id]["golden_text"]   = gcard.get("text", "")
            bg_db[base_id]["has_golden"]    = True

    # ── Aggregate personal game stats per card ──
    mode = request.args.get("mode", "all")
    conditions = ["placement > 0"]
    params: list = []
    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)
    where = "WHERE " + " AND ".join(conditions)

    personal: dict = {}   # base_card_id -> stats
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT final_board, placement FROM games {where}", params
        ).fetchall()

    for row in rows:
        placement = row["placement"]
        try:
            board = json.loads(row["final_board"] or "[]")
        except Exception:
            board = []
        for m in board:
            cid = m.get("card_id", "")
            if not cid:
                continue
            base = re.sub(r"_G$", "", cid)
            is_golden = m.get("golden", False) or cid.endswith("_G")
            if base not in personal:
                personal[base] = {"count": 0, "placements": [],
                                  "golden_count": 0, "top4": 0,
                                  "best": 99}
            e = personal[base]
            e["count"] += 1
            e["placements"].append(placement)
            if placement < e["best"]: e["best"] = placement
            if is_golden: e["golden_count"] += 1
            if placement <= 4: e["top4"] += 1

    # ── Build result (only base cards, golden merged in) ──
    result = []
    for cid, card in bg_db.items():
        p = personal.get(cid, {})
        cnt = p.get("count", 0)
        avg_p = round(sum(p["placements"]) / cnt, 2) if cnt else None
        races = card.get("races", ["NONE"])
        if isinstance(races, str):
            races = [races]
        has_golden = card.get("has_golden", False)
        in_pool = card.get("in_pool", False)
        # Categorise: token/generated cards end in t, t2, or are not in pool
        if in_pool:
            category = "一般"
        elif re.search(r"t\d*$", cid):
            category = "手下生成"
        else:
            category = "已移除"
        result.append({
            "card_id":       cid,
            "name":          card["name"],
            "attack":        card["attack"],
            "health":        card["health"],
            "tech_level":    card["tech_level"],
            "races":         races,
            "text":          _clean_card_text(card.get("text", "")),
            "golden_attack": card.get("golden_attack", 0),
            "golden_health": card.get("golden_health", 0),
            "golden_text":   _clean_card_text(card.get("golden_text", "")) if has_golden else "",
            "has_golden":    has_golden,
            "category":      category,
            # personal stats
            "seen":          cnt,
            "avg_placement": avg_p,
            "best":          p.get("best", None) if cnt else None,
            "golden_count":  p.get("golden_count", 0),
            "top4":          p.get("top4", 0),
            "top4_rate":     round(p["top4"] / cnt * 100, 1) if cnt else None,
        })

    # Apply category filter
    cat = request.args.get("cat", "")
    if cat:
        result = [m for m in result if m.get("category") == cat]

    result.sort(key=lambda x: (x["tech_level"], x["name"]))
    return jsonify({"minions": result, "total": len(result)})


@app.route("/api/delete-game/<game_id>", methods=["DELETE"])
def api_delete_game(game_id):
    with get_db() as conn:
        conn.execute("DELETE FROM games WHERE game_id = ?", (game_id,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/clear-db", methods=["DELETE"])
def api_clear_db():
    with get_db() as conn:
        deleted = conn.execute("SELECT COUNT(*) FROM games").fetchone()[0]
        conn.execute("DELETE FROM games")
        conn.commit()
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/api/top-builds")
def api_top_builds():
    mode = request.args.get("mode", "all")
    conditions = ["hero_name != ''"]
    params = []
    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)
    where = "WHERE " + " AND ".join(conditions)

    with get_db() as conn:
        # ── 英雄統計 ──
        hero_rows = conn.execute(f"""
            SELECT hero_name, hero_card_id,
                   COUNT(*) as games,
                   SUM(CASE WHEN placement=1 THEN 1 ELSE 0 END) as wins,
                   SUM(CASE WHEN placement<=4 THEN 1 ELSE 0 END) as top4,
                   AVG(placement) as avg_place
            FROM games {where}
            GROUP BY hero_name, hero_card_id
            ORDER BY avg_place ASC
        """, params).fetchall()

        # ── 讀所有場次（Python 側計算技能/飾品/隨從統計）──
        all_rows = conn.execute(
            f"SELECT hero_power_names, hero_power_ids, trinket_names, trinket_ids, "
            f"final_board, placement FROM games {where}", params
        ).fetchall()

    heroes = []
    for r in hero_rows:
        g = r["games"]
        heroes.append({
            "name":      r["hero_name"],
            "card_id":   r["hero_card_id"] or "",
            "games":     g,
            "wins":      r["wins"],
            "top4":      r["top4"],
            "win_rate":  round(r["wins"] / g * 100, 1) if g else 0,
            "top4_rate": round(r["top4"] / g * 100, 1) if g else 0,
            "avg_place": round(r["avg_place"], 2) if r["avg_place"] else 0,
        })

    # ── 技能統計 ──
    power_map: dict = {}
    trinket_map: dict = {}
    minion_map: dict = {}

    for row in all_rows:
        p = row["placement"]
        names = json.loads(row["hero_power_names"] or "[]")
        ids   = json.loads(row["hero_power_ids"]   or "[]")
        for i, name in enumerate(names):
            if not name:
                continue
            cid = ids[i] if i < len(ids) else ""
            if name not in power_map:
                power_map[name] = {"name": name, "card_id": cid, "games": 0, "wins": 0, "top4": 0, "total_p": 0}
            power_map[name]["games"]   += 1
            power_map[name]["total_p"] += p
            if p == 1: power_map[name]["wins"]  += 1
            if p <= 4: power_map[name]["top4"]  += 1

        tnames = json.loads(row["trinket_names"] or "[]")
        tids   = json.loads(row["trinket_ids"]   or "[]")
        for i, name in enumerate(tnames):
            if not name:
                continue
            cid = tids[i] if i < len(tids) else ""
            if name not in trinket_map:
                trinket_map[name] = {"name": name, "card_id": cid, "games": 0, "wins": 0, "top4": 0, "total_p": 0}
            trinket_map[name]["games"]   += 1
            trinket_map[name]["total_p"] += p
            if p == 1: trinket_map[name]["wins"]  += 1
            if p <= 4: trinket_map[name]["top4"]  += 1

        # 只統計第一名板面的隨從
        if p == 1:
            board = json.loads(row["final_board"] or "[]")
            for m in board:
                name = m.get("name", "")
                cid  = m.get("card_id", "")
                if not name:
                    continue
                if name not in minion_map:
                    minion_map[name] = {"name": name, "card_id": cid, "count": 0}
                minion_map[name]["count"] += 1

    def _fmt(d: dict) -> dict:
        g = d["games"]
        return {**d,
                "win_rate":  round(d["wins"] / g * 100, 1) if g else 0,
                "top4_rate": round(d["top4"] / g * 100, 1) if g else 0,
                "avg_place": round(d["total_p"] / g, 2)    if g else 0}

    powers   = sorted([_fmt(v) for v in power_map.values()   if v["games"] >= 2],
                      key=lambda x: x["avg_place"])
    trinkets = sorted([_fmt(v) for v in trinket_map.values() if v["games"] >= 2],
                      key=lambda x: x["avg_place"])
    minions  = sorted(minion_map.values(), key=lambda x: x["count"], reverse=True)[:20]

    return jsonify({"heroes": heroes, "powers": powers,
                    "trinkets": trinkets, "minions": minions})


@app.route("/tier-list")
def tier_list_page():
    return render_template("tier_list.html")


@app.route("/admin")
def admin_page():
    token = request.args.get("token", "")
    if not ADMIN_TOKEN or token != ADMIN_TOKEN:
        return "403 Forbidden", 403
    return render_template("admin.html", token=token)


@app.route("/api/spells")
def api_spells():
    if not os.path.exists(BG_SPELLS_CACHE):
        return jsonify({"spells": [], "total": 0})
    with open(BG_SPELLS_CACHE, encoding="utf-8") as f:
        spells = json.load(f)
    q = (request.args.get("q") or "").strip().lower()
    if q:
        spells = [s for s in spells if q in s.get("name", "").lower() or q in s.get("id", "").lower()]
    tier = request.args.get("tier")
    if tier:
        try:
            spells = [s for s in spells if s.get("tech_level") == int(tier)]
        except ValueError:
            pass
    # Assign category and clean text
    for s in spells:
        s["text"] = _clean_card_text(s.get("text") or "")
        sid = s.get("id", "")
        in_pool = s.get("in_pool", False)
        timewarp = s.get("timewarp", False)
        related = s.get("related_card")
        if sid.startswith("BGS_Treasures"):
            s["category"] = "暗月獎品"
        elif sid.startswith("BGDUO"):
            s["category"] = "雙人模式"
        elif "HeroPowerSpell" in sid:
            # Spells generated by legendary minions (e.g. 奧拉基爾之力)
            s["category"] = "手下產生"
        elif sid.startswith("BG34_Treasure") or timewarp:
            s["category"] = "時光扭曲"
        elif "戰巡艦" in (s.get("text") or ""):
            # Battlecruiser upgrade spells (generated by 索林姆's mechanic)
            s["category"] = "戰巡艦"
        elif in_pool:
            s["category"] = "一般"
        elif related:
            s["category"] = "手下產生"
        else:
            s["category"] = "已移除"
    cat = request.args.get("cat", "")
    if cat:
        spells = [s for s in spells if s.get("category") == cat]
    return jsonify({"spells": spells, "total": len(spells)})


@app.route("/api/cards")
def api_cards():
    """Combined minions + spells lookup by card ID or name."""
    q = (request.args.get("q") or "").strip().lower()
    results = []
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            minions = json.load(f)
        if q:
            minions = [m for m in minions if q in m.get("name", "").lower() or q in m.get("id", "").lower()]
        for m in minions:
            results.append({
                "id": m["id"], "name": m.get("name", ""), "type": "minion",
                "tech_level": m.get("tech_level", 0), "races": m.get("races", []),
                "text": m.get("text", "")
            })
    if os.path.exists(BG_SPELLS_CACHE):
        with open(BG_SPELLS_CACHE, encoding="utf-8") as f:
            spells = json.load(f)
        if q:
            spells = [s for s in spells if q in s.get("name", "").lower() or q in s.get("id", "").lower()]
        for s in spells:
            results.append({
                "id": s["id"], "name": s.get("name", ""), "type": "spell",
                "tech_level": s.get("tech_level", 0), "races": [],
                "text": s.get("text", "")
            })
    results.sort(key=lambda x: (x["tech_level"], x["name"]))
    return jsonify({"cards": results, "total": len(results)})


@app.route("/api/config")
def api_config():
    if not os.path.exists(BG_CONFIG_PATH):
        return jsonify({"season": 13, "hdt_build": "", "last_updated": ""})
    with open(BG_CONFIG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)
    return jsonify(cfg)


@app.route("/api/check-version")
def api_check_version():
    HDT_LOG = r"C:\Users\User\AppData\Roaming\HearthstoneDeckTracker\Logs\hdt_log.txt"
    import re as _re
    current_build = None
    if os.path.exists(HDT_LOG):
        try:
            with open(HDT_LOG, encoding="utf-8", errors="ignore") as f:
                content = f.read()
            matches = _re.findall(r"Build=(\d+)", content)
            if matches:
                current_build = matches[-1]
        except Exception:
            pass

    if current_build is None:
        return jsonify({"current_build": None, "needs_update": False})

    stored_build = ""
    season = 13
    if os.path.exists(BG_CONFIG_PATH):
        try:
            with open(BG_CONFIG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
            stored_build = cfg.get("hdt_build", "")
            season = cfg.get("season", 13)
        except Exception:
            pass

    needs_update = bool(stored_build) and current_build != stored_build
    return jsonify({
        "current_build": current_build,
        "stored_build": stored_build,
        "needs_update": needs_update,
        "season": season,
    })


def _git_push_data():
    """Commit changed data/ files and push to GitHub via GitHub API (no git CLI needed)."""
    import datetime as _dt, base64, json as _json
    import urllib.request as _urlreq, urllib.error as _urlerr

    pat = os.environ.get("GITHUB_PAT", "")
    if not pat:
        return {"pushed": False, "reason": "GITHUB_PAT not set"}

    repo = "insanehmt/HS_BG_Web"
    data_dir = os.path.join(os.path.dirname(__file__), "..", "data")
    api_base = f"https://api.github.com/repos/{repo}/contents"
    headers = {
        "Authorization": f"token {pat}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
        "User-Agent": "HS-BG-App",
    }

    def gh_get(path):
        req = _urlreq.Request(f"{api_base}/{path}", headers=headers)
        with _urlreq.urlopen(req, timeout=30) as r:
            return _json.loads(r.read())

    def gh_put(path, content_bytes, sha, message):
        body = _json.dumps({
            "message": message,
            "content": base64.b64encode(content_bytes).decode(),
            "sha": sha,
        }).encode()
        req = _urlreq.Request(f"{api_base}/{path}", data=body, headers=headers, method="PUT")
        with _urlreq.urlopen(req, timeout=60) as r:
            return _json.loads(r.read())

    # Files to sync
    data_files = ["bg_minions_cache.json", "bg_spells_cache.json",
                  "bg_trinkets_cache.json", "bg_heroes_cache.json", "bg_config.json"]

    now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    pushed = []
    skipped = []

    for fname in data_files:
        local_path = os.path.join(data_dir, fname)
        if not os.path.exists(local_path):
            continue
        with open(local_path, "rb") as f:
            local_bytes = f.read()

        try:
            remote = gh_get(f"data/{fname}")
            remote_content = base64.b64decode(remote["content"].replace("\n", ""))
            if local_bytes == remote_content:
                skipped.append(fname)
                continue
            try:
                gh_put(f"data/{fname}", local_bytes, remote["sha"],
                       f"auto: 更新排庫 {now} - {fname}")
                pushed.append(fname)
            except Exception as put_err:
                skipped.append(f"{fname}(put_err: {put_err})")
        except _urlerr.HTTPError as e:
            if e.code == 404:
                # File doesn't exist yet — create it (sha not needed for new files)
                try:
                    body = _json.dumps({
                        "message": f"auto: 新增 {fname} {now}",
                        "content": base64.b64encode(local_bytes).decode(),
                    }).encode()
                    req = _urlreq.Request(f"{api_base}/data/{fname}", data=body,
                                          headers=headers, method="PUT")
                    with _urlreq.urlopen(req, timeout=60):
                        pass
                    pushed.append(fname)
                except Exception as create_err:
                    skipped.append(f"{fname}(create_err: {create_err})")
            else:
                skipped.append(f"{fname}(err{e.code})")
        except Exception as e:
            skipped.append(f"{fname}(err: {e})")

    if not pushed:
        return {"pushed": False, "reason": "no changes", "skipped": skipped}

    return {
        "pushed": True,
        "files": pushed,
        "skipped": skipped,
    }


@app.route("/api/update-cards", methods=["POST"])
def api_update_cards():
    token = request.headers.get("X-Admin-Token", "")
    if not ADMIN_TOKEN or token != ADMIN_TOKEN:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    import subprocess, re as _re, datetime as _dt
    raw_path = os.path.join(os.path.dirname(__file__), "..", "data", "cards_zhtw_raw.json")

    # Download latest cards
    try:
        import urllib.request
        url = "https://api.hearthstonejson.com/v1/latest/zhTW/cards.json"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=120) as resp:
            with open(raw_path, "wb") as f:
                f.write(resp.read())
    except Exception as e:
        return jsonify({"success": False, "error": f"下載失敗: {e}"})

    try:
        with open(raw_path, encoding="utf-8") as f:
            all_cards = json.load(f)
    except Exception as e:
        return jsonify({"success": False, "error": f"解析下載檔失敗: {e}"})

    # --- Minions ---
    minions = []
    seen = set()
    for card in all_cards:
        cid = card.get("id", "")
        if not ((cid.startswith("BG") or cid.startswith("BGS_")) and
                card.get("type") == "MINION" and
                card.get("techLevel") and
                card.get("name") and
                cid not in seen):
            continue
        minions.append({
            "id": cid,
            "name": card.get("name", ""),
            "text": card.get("text", "").replace("\n", " "),
            "tech_level": card.get("techLevel", 0),
            "races": card.get("races", []),
            "attack": card.get("attack", 0),
            "health": card.get("health", 0),
            "mechanics": card.get("mechanics", []),
            "in_pool": bool(card.get("isBattlegroundsPoolMinion")),
        })
        seen.add(cid)
    minions.sort(key=lambda x: (x["tech_level"], x["name"]))
    with open(BG_MINIONS_CACHE, "w", encoding="utf-8") as f:
        json.dump(minions, f, ensure_ascii=False, indent=2)

    # --- Spells ---
    spells = []
    seen = set()
    for card in all_cards:
        cid = card.get("id", "")
        ctype = card.get("type", "")
        if not (cid.startswith("BG") and
                ctype in ("SPELL", "BATTLEGROUND_SPELL") and
                card.get("techLevel") and
                card.get("name") and
                cid not in seen):
            continue
        spells.append({
            "id": cid,
            "dbf_id": card.get("dbfId"),
            "name": card.get("name", ""),
            "text": card.get("text", "").replace("\n", " "),
            "type": ctype,
            "tech_level": card.get("techLevel", 0),
            "set": card.get("set", ""),
            "cost": card.get("cost", 0),
            "in_pool": bool(card.get("isBattlegroundsPoolSpell")),
            "timewarp": bool(card.get("battlegroundsTimewarpCard")),
            "related_card": card.get("battlegroundsRelatedCard"),
        })
        seen.add(cid)
    spells.sort(key=lambda x: (x["tech_level"], x["id"]))
    with open(BG_SPELLS_CACHE, "w", encoding="utf-8") as f:
        json.dump(spells, f, ensure_ascii=False, indent=2)

    # --- Heroes ---
    heroes = []
    seen = set()
    for card in all_cards:
        cid = card.get("id", "")
        if not (card.get("type") == "HERO" and
                cid.startswith("BG") and
                card.get("name") and
                cid not in seen):
            continue
        heroes.append({
            "id": cid,
            "name": card.get("name", ""),
            "text": card.get("text", "").replace("\n", " "),
        })
        seen.add(cid)
    with open(BG_HEROES_CACHE, "w", encoding="utf-8") as f:
        json.dump(heroes, f, ensure_ascii=False, indent=2)
    hero_count = len(heroes)

    # --- Trinkets ---
    trinkets = []
    seen = set()
    for card in all_cards:
        cid = card.get("id", "")
        ctype = card.get("type", "")
        mechanics = card.get("mechanics", [])
        if not ((ctype == "BATTLEGROUND_TRINKET" or
                 (cid.startswith("BG") and "TRINKET" in mechanics)) and
                card.get("name") and
                cid not in seen):
            continue
        trinkets.append({
            "id": cid,
            "name": card.get("name", ""),
            "text": card.get("text", "").replace("\n", " "),
            "tech_level": card.get("techLevel", 0),
            "tier": card.get("tier", ""),
        })
        seen.add(cid)
    if trinkets:
        with open(BG_TRINKETS_CACHE, "w", encoding="utf-8") as f:
            json.dump(trinkets, f, ensure_ascii=False, indent=2)

    # Clean up raw file
    try:
        os.remove(raw_path)
    except Exception:
        pass

    # Detect current build from HDT log
    HDT_LOG = r"C:\Users\User\AppData\Roaming\HearthstoneDeckTracker\Logs\hdt_log.txt"
    current_build = ""
    if os.path.exists(HDT_LOG):
        try:
            with open(HDT_LOG, encoding="utf-8", errors="ignore") as f:
                content = f.read()
            matches = _re.findall(r"Build=(\d+)", content)
            if matches:
                current_build = matches[-1]
        except Exception:
            pass

    # Update config
    cfg = {"season": 13, "hdt_build": "", "last_updated": ""}
    if os.path.exists(BG_CONFIG_PATH):
        try:
            with open(BG_CONFIG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception:
            pass
    cfg["hdt_build"] = current_build
    cfg["last_updated"] = _dt.datetime.now().isoformat()
    with open(BG_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

    # Auto-commit and push to GitHub if GITHUB_PAT is set
    try:
        git_result = _git_push_data()
    except Exception as e:
        git_result = {"pushed": False, "reason": f"git error: {e}"}

    return jsonify({
        "success": True,
        "updated": {
            "minions": len(minions),
            "spells": len(spells),
            "trinkets": len(trinkets),
            "heroes": hero_count,
        },
        "build": current_build,
        "git": git_result,
    })


def _get_valid_card_ids():
    """Return set of all card IDs in current minions + spells cache."""
    ids = set()
    for path in (BG_MINIONS_CACHE, BG_SPELLS_CACHE):
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                for card in json.load(f):
                    cid = card.get("id")
                    if cid:
                        ids.add(cid)
    return ids


def _annotate_comp_rotated(comp, valid_ids):
    """Return comp dict with added 'rotated' and 'rotated_cards' fields (in-memory only)."""
    if not valid_ids:
        return {**comp, "rotated": False, "rotated_cards": []}
    core_ids = comp.get("core", [])
    missing = [cid for cid in core_ids if cid and cid not in valid_ids]
    return {**comp, "rotated": bool(missing), "rotated_cards": missing}


@app.route("/api/tier-list")
def api_tier_list():
    if not os.path.exists(BG_COMPS_CACHE):
        return jsonify({"comps": [], "total": 0, "hidden": 0})
    with open(BG_COMPS_CACHE, encoding="utf-8") as f:
        comps = json.load(f)

    include_hidden = request.args.get("include_hidden") == "1"
    valid_ids = _get_valid_card_ids()

    annotated = [_annotate_comp_rotated(c, valid_ids) for c in comps]
    hidden_count = sum(1 for c in annotated if c["rotated"])

    if not include_hidden:
        visible = [c for c in annotated if not c["rotated"]]
    else:
        visible = annotated

    return jsonify({"comps": visible, "total": len(visible), "hidden": hidden_count})


@app.route("/api/tier-list/add", methods=["POST"])
def api_tier_list_add():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": "缺少組合名稱"}), 400
    if not data.get("races"):
        return jsonify({"success": False, "error": "缺少族裔"}), 400
    if not data.get("core"):
        return jsonify({"success": False, "error": "缺少核心卡牌"}), 400

    # Generate safe id from name
    import re, time
    slug = re.sub(r"[^\w]", "_", name)[:30].lower()
    comp_id = f"custom_{slug}_{int(time.time()) % 100000}"

    new_comp = {
        "id": comp_id,
        "tier": data.get("tier", "B"),
        "name": name,
        "races": data.get("races", []),
        "difficulty": data.get("difficulty", "medium"),
        "core": data.get("core", []),
        "core_names": data.get("core_names", []),
        "addon": data.get("addon", []),
        "addon_names": data.get("addon_names", []),
        "strategy": (data.get("strategy") or "").strip(),
        "tips": data.get("tips", []),
        "user_locked": True,   # 手動新增的 comp，爬蟲不覆蓋卡牌
    }

    if os.path.exists(BG_COMPS_CACHE):
        with open(BG_COMPS_CACHE, encoding="utf-8") as f:
            comps = json.load(f)
    else:
        comps = []

    comps.append(new_comp)
    with open(BG_COMPS_CACHE, "w", encoding="utf-8") as f:
        json.dump(comps, f, ensure_ascii=False, indent=2)

    return jsonify({"success": True, "comps": comps, "total": len(comps)})


@app.route("/api/tier-list/<comp_id>", methods=["PUT"])
def api_tier_list_edit(comp_id):
    data = request.get_json(silent=True) or {}
    comps = _load_comps()
    idx = next((i for i, c in enumerate(comps) if c.get("id") == comp_id), None)
    if idx is None:
        return jsonify({"success": False, "error": f"找不到 id={comp_id}"}), 404

    comp = comps[idx]
    card_fields = {"core", "core_names", "addon", "addon_names"}
    for field in ("name", "tier", "races", "difficulty", "core", "core_names",
                  "addon", "addon_names", "strategy", "tips"):
        if field in data:
            comp[field] = data[field]
    # 若用戶修改了卡牌，標記為手動編輯（爬蟲不覆蓋）
    if any(f in data for f in card_fields):
        comp["user_locked"] = True
    comps[idx] = comp
    _save_comps(comps)

    valid_ids = _get_valid_card_ids()
    annotated = [_annotate_comp_rotated(c, valid_ids) for c in comps]
    hidden_count = sum(1 for c in annotated if c["rotated"])
    return jsonify({"success": True, "comps": annotated, "total": len(annotated), "hidden": hidden_count})


# ---------------------------------------------------------------------------
# Scrape Firestone comps
# ---------------------------------------------------------------------------

@app.route("/api/scrape-comps", methods=["POST"])
def api_scrape_comps():
    """用 Playwright 從 Firestone 抓取最新牌組數據並更新 bg_comps.json。"""
    token = request.headers.get("X-Admin-Token", "")
    if not ADMIN_TOKEN or token != ADMIN_TOKEN:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    import sys, gzip as gz, traceback

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return jsonify({"success": False, "error": "未安裝 playwright，請執行 pip install playwright && playwright install chromium"}), 500

    # ── 讀取現有牌組（保留中文策略）──
    existing_map = {}
    if os.path.exists(BG_COMPS_CACHE):
        with open(BG_COMPS_CACHE, encoding="utf-8") as f:
            for c in json.load(f):
                existing_map[c["id"]] = c

    # ── 建立卡片中文名映射 ──
    name_map = {}
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            for m in json.load(f):
                cid = m.get("id") or m.get("card_id", "")
                nm  = m.get("name", "")
                if cid and nm:
                    name_map[cid] = nm

    # ── 族群映射 ──
    RACES_MAP = {
        'beast_self_damage': ['BEAST'], 'beast_stegodon': ['BEAST'],
        'demon_fodder': ['DEMON'], 'dragon_kalecgos': ['DRAGON'],
        'elemental_tier2_ballers': ['ELEMENTAL'],
        'quilboar_avenge': ['QUILBOAR'], 'quilboar_smuggler': ['QUILBOAR'],
        'mech_automaton': ['MECHANICAL'], 'mech_shield': ['MECHANICAL'],
        'murloc_mrrglton': ['MURLOC'], 'murloc_handbuff': ['MURLOC'],
        'murloc_scam': ['MURLOC'], 'naga_spellspam': ['NAGA'],
        'naga_deep_blue': ['NAGA'], 'pirate_bounty': ['PIRATE'],
        'undead_attack': ['UNDEAD'], 'undead_end_of_turn': ['UNDEAD'],
        'undead_overflow': ['UNDEAD'], 'neutral_back_to_back': [],
        'beast_banana': ['BEAST'], 'dragon_ring_bearer': ['DRAGON', 'NAGA'],
        'elemental_shop_buff': ['ELEMENTAL'],
    }
    NAME_ZH = {
        'beast_self_damage': '自傷野獸流', 'beast_stegodon': '劍龍野獸流',
        'demon_fodder': '食料惡魔流', 'dragon_kalecgos': '卡雷苟斯龍族流',
        'elemental_tier2_ballers': '二費球手元素流',
        'quilboar_avenge': '復仇野豬人流', 'quilboar_smuggler': '走私野豬人流',
        'mech_automaton': '自動機機械流', 'mech_shield': '護盾機械流',
        'murloc_mrrglton': '魚人鎮流', 'murloc_handbuff': '手牌強化魚人流',
        'murloc_scam': '詐騙魚人流', 'naga_spellspam': '法術狂納迦流',
        'naga_deep_blue': '深藍納迦流', 'pirate_bounty': '賞金海盜流',
        'undead_attack': '進攻亡靈流', 'undead_end_of_turn': '回合結束亡靈流',
        'undead_overflow': '溢出亡靈流', 'neutral_back_to_back': '連續觸發流',
        'beast_banana': '香蕉野獸流', 'dragon_ring_bearer': '戒指持有者龍族流',
        'elemental_shop_buff': '旅店強化元素流',
    }

    # ── 直接 HTTP 抓取（嘗試已知靜態端點）──
    DIRECT_URLS = [
        "https://static.firestoneapp.com/data/bgs-comps-strategies.json",
        "https://static.firestoneapp.com/data/bgs/bgs-comps-strategies.json",
        "https://static.firestoneapp.com/bgs-comps-strategies.json",
    ]

    def _try_direct_fetch():
        import urllib.request as _ur
        for url in DIRECT_URLS:
            try:
                req = _ur.Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept-Encoding": "gzip"})
                with _ur.urlopen(req, timeout=20) as r:
                    raw = r.read()
                try:
                    raw = gz.decompress(raw)
                except Exception:
                    pass
                data = json.loads(raw.decode("utf-8", "ignore"))
                if isinstance(data, list) and data and "compId" in data[0]:
                    return data
                if isinstance(data, dict) and isinstance(data.get("strategies"), list):
                    return data["strategies"]
            except Exception:
                pass
        return None

    # ── Playwright 抓取 ──
    strategies = None
    stats_map = {}

    # 先嘗試直接 HTTP 抓取（速度快、不需要 Playwright）
    strategies = _try_direct_fetch()

    if not strategies:
        try:
            with sync_playwright() as pw:
                browser = pw.chromium.launch(headless=True)
                page = browser.new_page()
                captured = {}        # url -> parsed JSON
                captured_urls = []   # 所有成功解析的 JSON URL（含非 firestone 域）

                def on_response(resp):
                    url = resp.url
                    # 略過靜態資源、圖片、字型
                    low = url.lower()
                    if any(low.endswith(ext) for ext in (
                        ".png", ".jpg", ".jpeg", ".webp", ".svg",
                        ".woff", ".woff2", ".ttf", ".otf", ".css",
                        ".js", ".map", ".ico",
                    )):
                        return
                    try:
                        body = resp.body()
                        if not body:
                            return
                        try:
                            body = gz.decompress(body)
                        except Exception:
                            pass
                        data = json.loads(body.decode("utf-8", "ignore"))
                        captured[url] = data
                        captured_urls.append(url)
                    except Exception:
                        pass

                page.on("response", on_response)
                try:
                    page.goto("https://www.firestoneapp.com/battlegrounds/comps",
                              wait_until="networkidle", timeout=60000)
                except Exception:
                    pass
                import time as _time
                _time.sleep(5)
                browser.close()

            # 根據內容結構辨識策略資料（不依賴域名或 URL 關鍵字）
            def _looks_like_strategies(data):
                if isinstance(data, list) and data and isinstance(data[0], dict):
                    first = data[0]
                    return "compId" in first or "cards" in first
                return False

            def _extract_comp_stats(data, stats_map):
                entries = []
                if isinstance(data, dict):
                    entries = data.get("compStats", [])
                elif isinstance(data, list):
                    entries = data
                for s in entries:
                    arch = s.get("archetype", "") or s.get("compId", "")
                    if arch:
                        stats_map[arch] = {
                            "avg_placement": round(s.get("averagePlacement", 0), 2),
                            "data_points":   s.get("dataPoints", 0),
                        }

            for url, data in captured.items():
                if _looks_like_strategies(data):
                    strategies = data
                    break
                if isinstance(data, dict):
                    inner = data.get("strategies") or data.get("comps") or data.get("data")
                    if isinstance(inner, list) and _looks_like_strategies(inner):
                        strategies = inner
                        break
                    if data.get("compStats"):
                        _extract_comp_stats(data, stats_map)

            # 第二輪：補充 comp-stats（strategies 找到後再掃一遍）
            if strategies:
                for url, data in captured.items():
                    if isinstance(data, dict) and data.get("compStats"):
                        _extract_comp_stats(data, stats_map)

            if not strategies:
                return jsonify({
                    "success": False,
                    "error": "無法取得 Firestone 牌組策略數據",
                    "captured_urls": captured_urls[:15],
                    "hint": "請把 captured_urls 貼給開發者，用於更新 API 端點",
                }), 500

        except Exception as e:
            return jsonify({"success": False, "error": f"Playwright 錯誤: {e}"}), 500

    if not strategies:
        return jsonify({"success": False, "error": "無法取得 Firestone 牌組策略數據"}), 500

    # ── 解析 ──
    comps = []
    for cd in strategies:
        comp_id = cd.get("compId", "")
        cards   = cd.get("cards", [])
        if not comp_id or not cards:
            continue

        core_ids  = [c["cardId"] for c in cards if c.get("status") == "CORE"]
        addon_ids = [c["cardId"] for c in cards if c.get("status") in ("ADDON", "CYCLE")]

        en_tips = [t["tip"] for t in cd.get("tips", [])
                   if t.get("language") == "enUS" and t.get("tip")]

        existing = existing_map.get(comp_id, {})
        strategy = existing.get("strategy", " ".join(en_tips[:2]) if en_tips else "")
        stat = stats_map.get(comp_id, {})

        comps.append({
            "id":            comp_id,
            "tier":          cd.get("powerLevel", "C"),
            "name":          existing.get("name") or NAME_ZH.get(comp_id, cd.get("name", comp_id)),
            "original_name": cd.get("name", comp_id),
            "races":         RACES_MAP.get(comp_id, existing.get("races", [])),
            "difficulty":    cd.get("difficulty", "Medium").lower(),
            "core":          core_ids,
            "core_names":    [name_map.get(c, c) for c in core_ids],
            "addon":         [addon_ids] if addon_ids else [],
            "addon_names":   [[name_map.get(c, c) for c in addon_ids]] if addon_ids else [],
            "strategy":      strategy,
            "tips":          existing.get("tips", []),
            "avg_placement": stat.get("avg_placement"),
            "data_points":   stat.get("data_points", 0),
            "patch":         cd.get("patchNumber"),
        })

    # ── 合併：原地更新，絕不刪除任何現有組合 ──
    # 規則：爬蟲只能新增或更新，舊組合全部保留。
    # Firestone compId → 舊 ID 對照（相同牌組，不同命名時期）
    FIRESTONE_ID_ALIASES = {
        "demon_fodder":            "fodder_demons",
        "quilboar_smuggler":       "smuggler_quilboar",
        "pirate_bounty":           "bounty_pirates",
        "neutral_back_to_back":    "back_to_back",
        "dragon_kalecgos":         "kalecgos_dragons",
        "quilboar_avenge":         "avenge_quilboar",
        "mech_automaton":          "automaton_mechs",
        "mech_shield":             "shield_mechs",
        "murloc_scam":             "scam_murlocs",
        "naga_spellspam":          "spellspam_nagas",
        "naga_deep_blue":          "deep_blue_nagas",
        "undead_attack":           "attack_undead",
        "beast_self_damage":       "self_damage_beasts",
        "beast_stegodon":          "stegodon_beasts",
        "elemental_tier2_ballers": "tier2_ballers",
        "murloc_mrrglton":         "mrrglton_murlocs",
        "murloc_handbuff":         "handbuff_murlocs",
        "undead_end_of_turn":      "end_of_turn_undead",
        "undead_overflow":         "overflow_undead",
        "beast_banana":            "beast_banana",
        "dragon_ring_bearer":      "dragon_ring_bearer",
        "elemental_shop_buff":     "elemental_shop_buff",
    }

    def _merge_cards(existing_ids, new_ids, nm):
        """合併卡片清單：保留現有順序，再追加新來源裡沒有的卡片。"""
        seen = set(existing_ids)
        merged = list(existing_ids)
        for cid in new_ids:
            if cid not in seen:
                merged.append(cid)
                seen.add(cid)
        return merged, [nm.get(c, c) for c in merged]

    def _apply_firestone(ec, comp):
        """將 Firestone 資料套用到現有 comp。
        - user_locked=True 的 comp：只更新 tier/stats，不動卡牌
        - 其他：直接用 Firestone 的卡牌（不再 union 累積）
        """
        ec["tier"]          = comp["tier"]
        ec["original_name"] = comp["original_name"]
        ec["difficulty"]    = comp["difficulty"]

        if not ec.get("user_locked"):
            # 直接使用 Firestone 的 core/addon（附上中文名補全）
            new_core, new_core_names = _merge_cards([], comp["core"], name_map)
            ec["core"]       = new_core
            ec["core_names"] = new_core_names
            if comp.get("addon"):
                merged_addons, merged_addon_names = [], []
                for grp in comp["addon"]:
                    mg, mn = _merge_cards([], grp, name_map)
                    merged_addons.append(mg)
                    merged_addon_names.append(mn)
                ec["addon"]       = merged_addons
                ec["addon_names"] = merged_addon_names

        ec["avg_placement"] = comp["avg_placement"]
        ec["data_points"]   = comp["data_points"]
        ec["patch"]         = comp["patch"]
        ec["name"]          = ec.get("name") or comp["name"]
        if not ec.get("strategy"):
            ec["strategy"] = comp["strategy"]
        if not ec.get("tips"):
            ec["tips"] = comp["tips"]
        return ec

    updated_count = 0
    added_count = 0

    for comp in comps:
        fid = comp["id"]
        # 1) 直接 ID 命中
        if fid in existing_map:
            existing_map[fid] = _apply_firestone(existing_map[fid], comp)
            updated_count += 1
        # 2) 有 alias → 原地更新舊 ID
        elif fid in FIRESTONE_ID_ALIASES:
            old_id = FIRESTONE_ID_ALIASES[fid]
            if old_id in existing_map:
                existing_map[old_id] = _apply_firestone(existing_map[old_id], comp)
                updated_count += 1
            else:
                # 舊 alias 也不存在 → 直接新增
                existing_map[fid] = comp
                added_count += 1
        # 3) 全新牌組 → 新增
        else:
            existing_map[fid] = comp
            added_count += 1

    # ── 所有 existing_map 的 comp 都保留（包含舊組合）──
    all_comps = list(existing_map.values())

    # 依 tier 排序
    tier_order = {"S": 0, "A": 1, "B": 2, "C": 3, "D": 4}
    all_comps.sort(key=lambda c: (tier_order.get(c["tier"], 9), c.get("name", "")))

    # ── 儲存 ──
    with open(BG_COMPS_CACHE, "w", encoding="utf-8") as f:
        json.dump(all_comps, f, ensure_ascii=False, indent=2)

    valid_ids = _get_valid_card_ids()
    annotated = [_annotate_comp_rotated(c, valid_ids) for c in all_comps]
    hidden_count = sum(1 for c in annotated if c["rotated"])

    return jsonify({
        "success":      True,
        "total":        len(all_comps),
        "firestone":    updated_count + added_count,
        "custom":       len(all_comps) - updated_count - added_count,
        "updated":      updated_count,
        "added":        added_count,
        "hidden":       hidden_count,
        "comps":        annotated,
        "stats_count":  len(stats_map),
        "patch":        all_comps[0].get("patch") if all_comps else None,
    })


# ---------------------------------------------------------------------------
# Scrape HSReplay comps
# ---------------------------------------------------------------------------

@app.route("/api/scrape-hsreplay", methods=["POST"])
def api_scrape_hsreplay():
    """用 Playwright 從 HSReplay 抓取牌組評級並更新 bg_comps.json。"""
    import sys, traceback, time as _time, re as _re

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return jsonify({"success": False, "error": "未安裝 playwright，請執行 pip install playwright && playwright install chromium"}), 500

    # 讀取現有牌組
    existing_map = {}
    if os.path.exists(BG_COMPS_CACHE):
        with open(BG_COMPS_CACHE, encoding="utf-8") as f:
            for c in json.load(f):
                existing_map[c["id"]] = c

    # 卡片名稱映射
    name_map = {}
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            for m in json.load(f):
                cid = m.get("id") or m.get("card_id", "")
                nm  = m.get("name", "")
                if cid and nm:
                    name_map[cid] = nm

    # HSReplay 英文名 → 我們的 comp ID
    HSREPLAY_NAME_TO_ID = {
        "Demons - Shop Buff":                  "demon_fodder",
        "Undead - Attack Scaling":             "undead_attack",
        "Back to Back":                        "neutral_back_to_back",
        "Pirates - Bounty APM":                "pirate_bounty",
        "Elementals - Shop Buff/Spells":       "elemental_shop_buff",
        "Mechs - Automaton":                   "mech_automaton",
        "Mechs - Automaton (SPECIFIC HEROES)": "mech_automaton",
        "Beasts - Leviathan":                  "beast_stegodon",
        "Beasts - Summons":                    "beast_self_damage",
        "Murlocs - Venom Scam":                "murloc_scam",
        "Murlocs - APM":                       "murloc_mrrglton",
        "Nagas - Combat Scaling/Spell Buff":   "naga_spellspam",
        "Dragons - Shiny Ring":                "dragon_ring_bearer",
        "Quilboar - Darkgaze":                 "quilboar_avenge",
        "Quilboar - Smuggler":                 "quilboar_smuggler",
        "Quilboar - Combat Scaling":           "quilboar_avenge",
        "Dragons - Spells":                    "dragon_kalecgos",
        "Nagas - Groundbreaker":               "naga_deep_blue",
        "Undead - Overflow":                   "undead_overflow",
        "Dragons - Battlecries":               "dragon_kalecgos",
    }

    # 族裔關鍵字推斷
    _RACE_KW = {
        "demon": "DEMON", "undead": "UNDEAD", "pirate": "PIRATE",
        "elemental": "ELEMENTAL", "mech": "MECHANICAL",
        "beast": "BEAST", "murloc": "MURLOC", "naga": "NAGA",
        "quilboar": "QUILBOAR", "dragon": "DRAGON",
    }

    JS_EXTRACT = """() => {
        var minLinks = Array.from(document.querySelectorAll('a[href*="/battlegrounds/minions/"]'));
        var seen = new Set();
        var comps = [];
        for (var link of minLinks) {
            var el = link;
            var compEl = null;
            for (var j = 0; j < 15; j++) {
                el = el.parentElement;
                if (!el) break;
                var cnt = el.querySelectorAll('a[href*="/battlegrounds/minions/"]').length;
                if (cnt >= 3 && cnt <= 20) { compEl = el; break; }
            }
            if (!compEl || seen.has(compEl)) continue;
            seen.add(compEl);
            var cards = Array.from(compEl.querySelectorAll('a[href*="/battlegrounds/minions/"]')).map(function(a) {
                var img = a.querySelector('img');
                var src = img ? img.src : '';
                var m = src.match(/\\/([A-Z][A-Z0-9_]+)\\.(png|webp)$/);
                return m ? m[1] : '';
            }).filter(Boolean);
            var nameEl = compEl;
            var compName = '', difficulty = '', tier = '';
            for (var k = 0; k < 12; k++) {
                nameEl = nameEl.parentElement;
                if (!nameEl) break;
                var text = nameEl.innerText || '';
                var lines = text.split('\\n').map(function(s){return s.trim();}).filter(Boolean);
                for (var line of lines) {
                    if ((line==='S'||line==='A'||line==='B'||line==='C') && !tier) tier=line;
                    if (/^(Medium|Easy|Hard)$/.test(line) && !difficulty) difficulty=line;
                    if (/[A-Za-z]{4,}/.test(line) && !/(Medium|Easy|Hard|hearthstone|Comps|Heroes|Guides|Season|Tier7|HSReplay|Social|Download|Copyright|navigation|Battlegrounds|Sign|Match|Last|Powered|JeefHS|Scale|Summon|Make|Buy|Spend|Play|Cast|Cycle|Overflow|Bounce|Stack|Build|Standard|Meta|Arena|Rank|English|Patch|Updated|Win|Core|Minions|Tier|Comp|Difficulty)/.test(line) && !compName && line.length>4 && line.length<65) {
                        compName = line;
                    }
                }
                if (compName && tier) break;
            }
            if (cards.length > 0) comps.push({name: compName, tier: tier, difficulty: difficulty.toLowerCase(), cards: cards});
        }
        return comps;
    }"""

    raw_comps = []
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 900},
            )
            page = ctx.new_page()
            try:
                page.goto("https://hsreplay.net/battlegrounds/comps/", wait_until="domcontentloaded", timeout=20000)
            except Exception:
                pass
            _time.sleep(4)
            # scroll to trigger lazy loading
            for y in range(0, 8000, 600):
                page.evaluate(f"window.scrollTo(0, {y})")
                _time.sleep(0.2)
            _time.sleep(1)
            raw_comps = page.evaluate(JS_EXTRACT)
            browser.close()
    except Exception as e:
        return jsonify({"success": False, "error": f"Playwright 錯誤：{e}\n{traceback.format_exc()}"}), 500

    if not raw_comps:
        return jsonify({"success": False, "error": "HSReplay 頁面未回傳牌組數據"}), 500

    # Firestone ID → 舊 ID alias（與 Firestone 爬蟲共用同一映射，避免重複新增）
    _HS_ALIASES = {
        "demon_fodder": "fodder_demons", "quilboar_smuggler": "smuggler_quilboar",
        "pirate_bounty": "bounty_pirates", "neutral_back_to_back": "back_to_back",
        "dragon_kalecgos": "kalecgos_dragons", "quilboar_avenge": "avenge_quilboar",
        "mech_automaton": "automaton_mechs", "mech_shield": "shield_mechs",
        "murloc_scam": "scam_murlocs", "naga_spellspam": "spellspam_nagas",
        "naga_deep_blue": "deep_blue_nagas", "undead_attack": "attack_undead",
        "beast_self_damage": "self_damage_beasts", "beast_stegodon": "stegodon_beasts",
        "elemental_tier2_ballers": "tier2_ballers", "murloc_mrrglton": "mrrglton_murlocs",
        "murloc_handbuff": "handbuff_murlocs", "undead_end_of_turn": "end_of_turn_undead",
        "undead_overflow": "overflow_undead",
    }

    updated_count = 0
    added_count = 0
    updated_ids = set()

    for hc in raw_comps:
        name  = hc.get("name", "").strip()
        tier  = hc.get("tier", "").strip()
        diff  = hc.get("difficulty", "").strip()
        cards = hc.get("cards", [])

        # 找對應的 comp ID
        comp_id = HSREPLAY_NAME_TO_ID.get(name)
        if comp_id:
            # 如果 comp_id 不在 existing_map，嘗試用 alias 找到舊 ID
            if comp_id not in existing_map and comp_id in _HS_ALIASES:
                old_id = _HS_ALIASES[comp_id]
                if old_id in existing_map:
                    comp_id = old_id
        else:
            # 嘗試 race 關鍵字匹配
            lname = name.lower()
            for kw, race in _RACE_KW.items():
                if kw in lname:
                    for eid, ec in existing_map.items():
                        if race in ec.get("races", []) and eid not in updated_ids:
                            comp_id = eid
                            break
                if comp_id:
                    break

        if not comp_id:
            # 跳過無法匹配的牌組
            continue

        updated_ids.add(comp_id)
        if comp_id in existing_map:
            ec = existing_map[comp_id]
            # HSReplay 不覆蓋 tier（Firestone 評分優先）；只補齊 difficulty，合併 core 卡片
            if diff and not ec.get("difficulty"):
                ec["difficulty"] = diff
            # core 卡片：user_locked 的 comp 不覆蓋；其他直接用 HSReplay 的
            if cards and not ec.get("user_locked"):
                ec["core"]       = cards
                ec["core_names"] = [name_map.get(cid, cid) for cid in cards]
            existing_map[comp_id] = ec
            updated_count += 1
        else:
            # 推斷族裔
            lname = name.lower()
            races = []
            for kw, race in _RACE_KW.items():
                if kw in lname and race not in races:
                    races.append(race)
            new_comp = {
                "id":           comp_id,
                "tier":         tier or "B",
                "name":         name,
                "original_name": name,
                "races":        races,
                "difficulty":   diff or "medium",
                "core":         cards,
                "core_names":   [name_map.get(cid, cid) for cid in cards],
                "addon":        [],
                "addon_names":  [],
                "strategy":     "",
                "tips":         [],
                "avg_placement": None,
                "data_points":  None,
                "patch":        None,
            }
            existing_map[comp_id] = new_comp
            added_count += 1

    all_comps = list(existing_map.values())
    kept = len(all_comps) - updated_count - added_count
    tier_order = {"S": 0, "A": 1, "B": 2, "C": 3, "D": 4}
    all_comps.sort(key=lambda c: (tier_order.get(c["tier"], 9), c["name"]))

    with open(BG_COMPS_CACHE, "w", encoding="utf-8") as f:
        json.dump(all_comps, f, ensure_ascii=False, indent=2)

    valid_ids = _get_valid_card_ids()
    annotated = [_annotate_comp_rotated(c, valid_ids) for c in all_comps]
    hidden_count = sum(1 for c in annotated if c["rotated"])

    return jsonify({
        "success":  True,
        "total":    len(all_comps),
        "updated":  updated_count,
        "added":    added_count,
        "kept":     kept,
        "hidden":   hidden_count,
        "comps":    annotated,
    })




# Race keyword → BG race tag mapping
_RACE_NAME_MAP = {
    "demon": "DEMON", "demons": "DEMON",
    "undead": "UNDEAD",
    "pirate": "PIRATE", "pirates": "PIRATE",
    "elemental": "ELEMENTAL", "elementals": "ELEMENTAL",
    "mech": "MECHANICAL", "mechs": "MECHANICAL", "mechanical": "MECHANICAL",
    "beast": "BEAST", "beasts": "BEAST",
    "murloc": "MURLOC", "murlocs": "MURLOC",
    "naga": "NAGA", "nagas": "NAGA",
    "quilboar": "QUILBOAR",
    "dragon": "DRAGON", "dragons": "DRAGON",
}


def _build_card_race_lookup():
    """Return {card_id: [races]} from minions cache."""
    lookup = {}
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            for card in json.load(f):
                cid = card.get("id")
                races = [r for r in (card.get("races") or []) if r and r != "NONE" and r != "ALL"]
                if cid and races:
                    lookup[cid] = races
    return lookup


def _infer_races(comp_name: str, core_ids: list) -> list:
    """Infer race tags from comp name keywords + core card races."""
    races = set()

    # 1. Parse from name (HSReplay format "Race - Strategy" or any keyword)
    name_lower = (comp_name or "").lower()
    for keyword, race in _RACE_NAME_MAP.items():
        if keyword in name_lower:
            races.add(race)

    # 2. Look up from core card IDs
    card_race_lookup = _build_card_race_lookup()
    for cid in (core_ids or []):
        for r in card_race_lookup.get(cid, []):
            if r not in ("NONE", "ALL"):
                races.add(r)

    return sorted(races)

def _build_auto_strategy(comp_name: str, description: str) -> str:
    """Auto-generate strategy text from HSReplay comp_name + description.
    comp_name format: "Race - Strategy" e.g. "Demons - Shop Buff"
    description: English one-liner e.g. "Scale and consume shop"
    """
    parts = []
    if " - " in (comp_name or ""):
        race_en, strat_en = comp_name.split(" - ", 1)
        parts.append(f"【{race_en.strip()} · {strat_en.strip()}】")
    if description:
        parts.append(description.strip())
    return " ".join(parts)


def _slugify(text):
    import re
    return re.sub(r"[^\w]+", "_", text.strip().lower()).strip("_")


def _load_comps():
    if os.path.exists(BG_COMPS_CACHE):
        with open(BG_COMPS_CACHE, encoding="utf-8") as f:
            return json.load(f)
    return []


def _save_comps(comps):
    with open(BG_COMPS_CACHE, "w", encoding="utf-8") as f:
        json.dump(comps, f, ensure_ascii=False, indent=2)


def _match_comp(parsed_name, comps):
    """Return (existing_comp or None) by id/name similarity."""
    pid = _slugify(parsed_name)
    plow = parsed_name.lower()
    # 1. Exact id match
    for c in comps:
        if c.get("id", "") == pid:
            return c
    # 2. Name substring match (use original_name if present)
    for c in comps:
        cname = (c.get("original_name") or c.get("name", "")).lower()
        ename = c.get("name", "").lower()
        if plow in cname or cname in plow or plow in ename or ename in plow:
            return c
    return None


@app.route("/api/parse-firestone", methods=["POST"])
def api_parse_firestone():
    import re
    data = request.get_json(silent=True) or {}
    text = data.get("text", "")
    if not text.strip():
        return jsonify({"error": "缺少 text"}), 400

    comps = _load_comps()
    blocks = re.split(r"\n{2,}", text.strip())
    parsed = []

    for block in blocks:
        lines = [l.strip() for l in block.strip().splitlines() if l.strip()]
        if not lines:
            continue

        games_line = None
        power = None
        position = None
        tier = None
        difficulty = None
        name = None

        games_re = re.compile(r"^\d[\d,]*\s+games$", re.IGNORECASE)
        float_re  = re.compile(r"^(\d+\.\d+)$")
        tier_re   = re.compile(r"^[SABC]$")
        diff_re   = re.compile(r"^(easy|medium|hard)$", re.IGNORECASE)

        for line in lines:
            if games_re.match(line):
                games_line = line
            elif tier_re.match(line):
                tier = line
            elif diff_re.match(line):
                difficulty = line.lower()
            elif float_re.match(line):
                val = float(line)
                if power is None:
                    power = val
                elif position is None:
                    position = val
            elif name is None:
                name = line

        if not name or not tier:
            continue

        games_str = ""
        if games_line:
            games_str = games_line.split()[0]

        comp_id = _slugify(name)
        existing = _match_comp(name, comps)
        action = "update" if existing else "new"

        parsed.append({
            "name": name,
            "tier": tier,
            "difficulty": difficulty or "medium",
            "games": games_str,
            "power": power,
            "position": position,
            "action": action,
            "existing_id": existing["id"] if existing else None,
            "existing_name": existing.get("name") if existing else None,
            "id": comp_id,
        })

    return jsonify({"parsed": parsed, "total": len(parsed)})


@app.route("/api/import-firestone", methods=["POST"])
def api_import_firestone():
    data = request.get_json(silent=True) or {}
    incoming = data.get("comps", [])
    if not isinstance(incoming, list):
        return jsonify({"error": "comps 必須是陣列"}), 400

    comps = _load_comps()
    updated = 0
    added = 0

    for item in incoming:
        action = item.get("action")
        if action == "update" and item.get("existing_id"):
            for c in comps:
                if c["id"] == item["existing_id"]:
                    c["tier"] = item.get("tier", c["tier"])
                    c["difficulty"] = item.get("difficulty", c["difficulty"])
                    updated += 1
                    break
        elif action == "new":
            comp_id = item.get("id") or _slugify(item.get("name", ""))
            inferred_races = _infer_races(item.get("name", ""), [])
            new_comp = {
                "id": comp_id,
                "tier": item.get("tier", "B"),
                "name": item.get("name", ""),
                "original_name": item.get("name", ""),
                "races": inferred_races,
                "difficulty": item.get("difficulty", "medium"),
                "core": [],
                "core_names": [],
                "addon": [],
                "addon_names": [],
                "strategy": "",
                "tips": [],
            }
            comps.append(new_comp)
            added += 1

    _save_comps(comps)
    return jsonify({"success": True, "updated": updated, "added": added,
                    "comps": comps, "total": len(comps)})


# ---------------------------------------------------------------------------
# Skill 3: HSReplay Paste Import
# ---------------------------------------------------------------------------

def _build_card_name_lookup():
    lookup = {}
    for cache_path in (BG_MINIONS_CACHE, BG_SPELLS_CACHE):
        if os.path.exists(cache_path):
            with open(cache_path, encoding="utf-8") as f:
                cards = json.load(f)
            for card in cards:
                name = card.get("name", "").strip()
                if name:
                    lookup[name] = card.get("id", "")
    return lookup


@app.route("/api/parse-hsreplay", methods=["POST"])
def api_parse_hsreplay():
    import re
    data = request.get_json(silent=True) or {}
    text = data.get("text", "")
    if not text.strip():
        return jsonify({"error": "缺少 text"}), 400

    lines = text.splitlines()

    # Find start: line AFTER "Core Minions"
    start_idx = None
    for i, line in enumerate(lines):
        if line.strip() == "Core Minions":
            start_idx = i + 1
            break
    if start_idx is None:
        return jsonify({"error": "找不到 'Core Minions' 段落"}), 400

    # Find end: "HSReplay" footer line (after start)
    end_idx = len(lines)
    for i in range(start_idx, len(lines)):
        l = lines[i].strip()
        if l in ("HSReplay", "HSReplay.net"):
            end_idx = i
            break

    data_lines = lines[start_idx:end_idx]
    data_text = "\n".join(data_lines)

    comps_existing = _load_comps()
    card_lookup = _build_card_name_lookup()
    diff_re = re.compile(r"^(easy|medium|hard)$", re.IGNORECASE)

    blocks = re.split(r"\n{2,}", data_text.strip())
    parsed = []

    for block in blocks:
        blines = [l.strip() for l in block.strip().splitlines() if l.strip()]
        if len(blines) < 4:
            continue
        # Line 0: header card (Chinese), Line 1: English comp name (must have " - ")
        header_card = blines[0]
        comp_name   = blines[1]
        if " - " not in comp_name:
            continue
        description = blines[2] if len(blines) > 2 else ""
        difficulty  = "medium"
        core_start  = 4
        if len(blines) > 3 and diff_re.match(blines[3]):
            difficulty = blines[3].lower()
        else:
            core_start = 3

        core_names_raw = [header_card] + blines[core_start:]
        # Deduplicate preserving order
        seen_names: set = set()
        core_names: list = []
        for cn in core_names_raw:
            if cn not in seen_names:
                seen_names.add(cn)
                core_names.append(cn)

        core_ids = [card_lookup.get(cn) or None for cn in core_names]

        # Generate id from part after " - " + race prefix
        parts = comp_name.split(" - ", 1)
        race_part = parts[0].strip().lower()
        type_part = parts[1].strip().lower() if len(parts) > 1 else ""
        comp_id = _slugify(type_part + "_" + race_part)

        existing = _match_comp(comp_name, comps_existing)
        action = "update" if existing else "new"
        # Auto-fill tier from existing data; new comps need manual selection
        auto_tier = existing.get("tier") if existing else None

        parsed.append({
            "comp_name": comp_name,
            "description": description,
            "difficulty": difficulty,
            "tier": auto_tier,
            "core_names": core_names,
            "core_ids": core_ids,
            "action": action,
            "existing_id": existing["id"] if existing else None,
            "existing_name": existing.get("name") if existing else None,
            "id": comp_id,
        })

    return jsonify({"parsed": parsed, "total": len(parsed)})


@app.route("/api/import-hsreplay", methods=["POST"])
def api_import_hsreplay():
    data = request.get_json(silent=True) or {}
    incoming = data.get("comps", [])
    if not isinstance(incoming, list):
        return jsonify({"error": "comps 必須是陣列"}), 400

    comps = _load_comps()
    updated = 0
    added = 0

    for item in incoming:
        action = item.get("action")
        core_names = item.get("core_names", [])
        core_ids_raw = item.get("core_ids", [])
        # Paired filter: keep entries where id is not null
        core_pairs = [(cid, cn) for cid, cn in zip(core_ids_raw, core_names) if cid]
        core_ids_clean = [p[0] for p in core_pairs]
        core_names_clean = [p[1] for p in core_pairs]

        if action == "update" and item.get("existing_id"):
            for c in comps:
                if c["id"] == item["existing_id"]:
                    c["difficulty"] = item.get("difficulty", c["difficulty"])
                    if item.get("tier"):
                        c["tier"] = item["tier"]
                    if core_ids_clean:
                        c["core"] = core_ids_clean
                        c["core_names"] = core_names_clean
                    # Always update hsreplay_desc if available
                    if item.get("description"):
                        c["hsreplay_desc"] = item["description"]
                    # Auto-fill races if missing
                    if not c.get("races"):
                        c["races"] = _infer_races(item.get("comp_name",""), core_ids_clean)
                    # Always update strategy from HSReplay if description is available
                    # (overwrite auto-generated bracket-only strategy too)
                    if item.get("description") or not c.get("strategy"):
                        c["strategy"] = _build_auto_strategy(
                            item.get("comp_name", ""), item.get("description", ""))
                    updated += 1
                    break
        elif action == "new":
            comp_id = item.get("id") or _slugify(item.get("comp_name", ""))
            inferred_races = _infer_races(item.get("comp_name", ""), core_ids_clean)
            new_comp = {
                "id": comp_id,
                "tier": item.get("tier") or "B",
                "name": item.get("comp_name", ""),
                "original_name": item.get("comp_name", ""),
                "hsreplay_desc": item.get("description", ""),
                "races": inferred_races,
                "difficulty": item.get("difficulty", "medium"),
                "core": core_ids_clean,
                "core_names": core_names_clean,
                "addon": [],
                "addon_names": [],
                "strategy": _build_auto_strategy(
                    item.get("comp_name", ""), item.get("description", "")),
                "tips": [],
            }
            comps.append(new_comp)
            added += 1

    _save_comps(comps)
    return jsonify({"success": True, "updated": updated, "added": added,
                    "comps": comps, "total": len(comps)})


@app.route("/api/fix-races", methods=["POST"])
def api_fix_races():
    """Back-fill race tags for all comps that have empty races field."""
    comps = _load_comps()
    fixed = 0
    for c in comps:
        if not c.get("races"):
            inferred = _infer_races(
                c.get("original_name") or c.get("name", ""),
                c.get("core", [])
            )
            if inferred:
                c["races"] = inferred
                fixed += 1
    _save_comps(comps)
    return jsonify({"success": True, "fixed": fixed, "total": len(comps)})


@app.route("/api/fix-strategy", methods=["POST"])
def api_fix_strategy():
    """Back-fill/upgrade strategy for HSReplay comps using original_name + hsreplay_desc."""
    comps = _load_comps()
    fixed = 0
    for c in comps:
        orig = c.get("original_name") or c.get("name", "")
        if " - " not in orig:
            continue
        desc = c.get("hsreplay_desc", "")
        new_strat = _build_auto_strategy(orig, desc)
        # Update if: empty, or bracket-only (no description appended yet but desc now available)
        current = c.get("strategy", "")
        if not current or (desc and current == _build_auto_strategy(orig, "")):
            c["strategy"] = new_strat
            fixed += 1
    _save_comps(comps)
    return jsonify({"success": True, "fixed": fixed, "total": len(comps)})


@app.route("/api/card-picker")
def api_card_picker():
    """Return cards for the visual card picker. type=minion|spell|trinket|hero_power"""
    card_type = request.args.get("type", "minion")
    q         = request.args.get("q", "").strip().lower()
    race      = request.args.get("race", "")

    cards = []

    if card_type == "minion":
        raw  = _load_json(BG_MINIONS_CACHE)
        seen = set()
        for c in raw:
            cid = c.get("id", "")
            if cid.endswith("_G") or cid in seen:
                continue
            seen.add(cid)
            card_races = c.get("races") or []
            if race and race != "ALL":
                if race not in card_races:
                    continue
            name = c.get("name", "")
            text = c.get("text", "") or ""
            if q and q not in name.lower() and q not in text.lower() and q not in cid.lower():
                continue
            cards.append({
                "id":    cid,
                "name":  name,
                "tl":    c.get("tech_level", 1),
                "races": card_races,
                "atk":   c.get("attack", 0),
                "hp":    c.get("health", 0),
                "text":  text,
            })
        cards.sort(key=lambda x: (x["tl"], x["name"]))

    elif card_type == "spell":
        raw = _load_json(BG_SPELLS_CACHE)
        for c in raw:
            cid  = c.get("id", "")
            if cid.endswith("_G"):
                continue
            name = c.get("name", "")
            text = c.get("text", "") or ""
            if q and q not in name.lower() and q not in text.lower() and q not in cid.lower():
                continue
            cards.append({
                "id":   cid,
                "name": name,
                "cost": c.get("cost", 0),
                "text": _clean_card_text(text),
            })
        cards.sort(key=lambda x: (x["cost"] or 0, x["name"]))

    elif card_type == "trinket":
        raw = _load_json(BG_TRINKETS_CACHE)
        for c in raw:
            cid  = c.get("id", "")
            if cid.endswith("_G"):
                continue
            name = c.get("name", "")
            text = c.get("text", "") or ""
            if q and q not in name.lower() and q not in text.lower() and q not in cid.lower():
                continue
            cards.append({"id": cid, "name": name, "text": _clean_card_text(text)})
        cards.sort(key=lambda x: x["name"])

    elif card_type == "hero_power":
        raw = _load_json(HS_CARDS_FULL) or {}
        for cid, v in raw.items():
            if not cid.startswith("BG") or v.get("type") != "HERO_POWER":
                continue
            name = v.get("name", "")
            text = v.get("text", "") or ""
            if q and q not in name.lower() and q not in text.lower() and q not in cid.lower():
                continue
            cards.append({"id": cid, "name": name, "text": _clean_card_text(text)})
        cards.sort(key=lambda x: x["name"])

    return jsonify({"cards": cards})



def _clean_card_text(text):
    """Clean Hearthstone card text for HTML display."""
    if not text:
        return ""
    # Remove HS markup tokens like [x], [$d2], [$a2], [b], etc.
    text = re.sub(r'\[[^\]]{0,8}\]', '', text)
    # Remove duplicate nested bold: <b><b>X</b></b> → <b>X</b>
    text = re.sub(r'<b><b>(.*?)</b></b>', r'<b>\1</b>', text)
    # Remove repeated quest stage text
    # Pattern: "(剩N回合！)</i>N..." → strip number and repeated content
    text = re.sub(r'(回合！\)(?:</i>)?)\d+.*', r'\1', text, flags=re.DOTALL)
    text = re.sub(r'(\(完成！\)(?:</i>)?).*', r'\1', text, flags=re.DOTALL)
    # Newlines → <br>
    text = text.replace('\n', '<br>')
    return text.strip()


def _load_json(path):
    """Load a JSON file with UTF-8 encoding."""
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# 資料同步端點：本地推送資料到雲端
# ---------------------------------------------------------------------------

# 允許同步的白名單檔案（防止任意覆寫）
_SYNC_ALLOWED = {
    "bg_comps.json",
    "bg_minions_cache.json",
    "bg_spells_cache.json",
    "bg_trinkets_cache.json",
    "bg_heroes_cache.json",
    "hero_meta.json",
    "hsreplay_meta_cache.json",
    "bg_config.json",
    "hs_bg_heroes.json",
}

_DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")


@app.route("/api/push-data", methods=["POST"])
def api_push_data():
    """接收本地端推送的 JSON 資料，更新雲端資料檔案。
    需要在環境變數設定 SYNC_TOKEN，請求 body: {"token": "...", "files": {"bg_comps.json": [...], ...}}
    """
    sync_token = os.environ.get("SYNC_TOKEN", "")
    if not sync_token:
        return jsonify({"success": False, "error": "伺服器未設定 SYNC_TOKEN"}), 503

    body = request.get_json(silent=True) or {}
    if body.get("token") != sync_token:
        return jsonify({"success": False, "error": "token 錯誤"}), 403

    files = body.get("files", {})
    if not files:
        return jsonify({"success": False, "error": "未包含任何檔案"}), 400

    # Guard against oversized payloads (each file capped at 10 MB serialised)
    _MAX_FILE_BYTES = 10 * 1024 * 1024
    updated = []
    skipped = []
    for fname, content in files.items():
        if fname not in _SYNC_ALLOWED:
            skipped.append(fname)
            continue
        serialised = json.dumps(content, ensure_ascii=False, indent=2)
        if len(serialised.encode()) > _MAX_FILE_BYTES:
            skipped.append(f"{fname}(too large)")
            continue
        target = os.path.join(_DATA_DIR, fname)
        os.makedirs(os.path.dirname(target), exist_ok=True)
        with open(target, "w", encoding="utf-8") as f:
            f.write(serialised)
        updated.append(fname)

    return jsonify({
        "success": True,
        "updated": len(updated),
        "files":   updated,
        "skipped": skipped,
    })


@app.route("/api/sync-to-cloud", methods=["POST"])
def api_sync_to_cloud():
    """從本地 Flask 推送資料到雲端（讀取 .push_config）。"""
    if PUBLIC_MODE:
        return jsonify({"success": False, "error": "雲端模式不支援同步"}), 403

    body = request.get_json(silent=True) or {}
    full = body.get("full", False)   # full=True → 包含大型卡牌快取

    # 讀取 .push_config
    config_path = os.path.join(os.path.dirname(__file__), "..", ".push_config")
    web_url, sync_token = "", ""
    if os.path.exists(config_path):
        for line in open(config_path, encoding="utf-8").read().splitlines():
            if line.startswith("WEB_URL="):
                web_url = line[len("WEB_URL="):]
            elif line.startswith("SYNC_TOKEN="):
                sync_token = line[len("SYNC_TOKEN="):]

    if not web_url or not sync_token:
        return jsonify({"success": False,
                        "error": "找不到 .push_config（需含 WEB_URL 和 SYNC_TOKEN）"}), 400

    default_files = [
        "bg_comps.json", "hero_meta.json",
        "hsreplay_meta_cache.json", "bg_config.json", "hs_bg_heroes.json",
    ]
    all_files = default_files + [
        "bg_minions_cache.json", "bg_spells_cache.json",
        "bg_trinkets_cache.json", "bg_heroes_cache.json",
    ]
    files_to_sync = all_files if full else default_files

    data_dir = os.path.join(os.path.dirname(__file__), "..", "data")
    payload = {}
    for fname in files_to_sync:
        fpath = os.path.join(data_dir, fname)
        if os.path.exists(fpath):
            with open(fpath, encoding="utf-8") as f:
                payload[fname] = json.load(f)

    if not payload:
        return jsonify({"success": False, "error": "沒有可同步的資料檔"}), 400

    try:
        import requests as req
        push_url = web_url.rstrip("/") + "/api/push-data"
        r = req.post(push_url, json={"token": sync_token, "files": payload}, timeout=60)
    except Exception as e:
        return jsonify({"success": False, "error": f"連線失敗：{e}"}), 500

    if r.status_code == 200:
        result = r.json()
        return jsonify({
            "success": True,
            "updated": result.get("updated", 0),
            "files":   result.get("files", []),
            "web_url": web_url,
        })
    elif r.status_code == 403:
        return jsonify({"success": False, "error": "SYNC_TOKEN 錯誤"}), 403
    else:
        return jsonify({"success": False,
                        "error": f"雲端回應 HTTP {r.status_code}"}), 500


@app.route("/heroes")
def heroes_page():
    return render_template("heroes.html")


@app.route("/api/heroes")
def api_heroes():
    mode = request.args.get("mode", "all")
    sort = request.args.get("sort", "avg_place")   # avg_place / win_rate / top4_rate / games

    conditions = ["placement > 0", "hero_name != ''"]
    params: list = []
    if mode in ("solo", "duo"):
        conditions.append("game_mode = ?")
        params.append(mode)
    where = "WHERE " + " AND ".join(conditions)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT hero_name, hero_card_id,
                   COUNT(*) as games,
                   SUM(CASE WHEN placement=1 THEN 1 ELSE 0 END) as wins,
                   SUM(CASE WHEN placement<=4 THEN 1 ELSE 0 END) as top4,
                   AVG(placement) as avg_place,
                   MIN(placement) as best,
                   MAX(placement) as worst
            FROM games {where}
            GROUP BY hero_name, hero_card_id
        """, params).fetchall()

    heroes = []
    for r in rows:
        g = r["games"]
        heroes.append({
            "name":       r["hero_name"],
            "card_id":    r["hero_card_id"] or "",
            "games":      g,
            "wins":       r["wins"],
            "top4":       r["top4"],
            "best":       r["best"],
            "worst":      r["worst"],
            "win_rate":   round(r["wins"] / g * 100, 1) if g else 0,
            "top4_rate":  round(r["top4"] / g * 100, 1) if g else 0,
            "avg_place":  round(r["avg_place"], 2) if r["avg_place"] else 0,
        })

    sort_key = {
        "avg_place":  lambda x: (x["avg_place"], -x["games"]),
        "win_rate":   lambda x: (-x["win_rate"], x["avg_place"]),
        "top4_rate":  lambda x: (-x["top4_rate"], x["avg_place"]),
        "games":      lambda x: (-x["games"], x["avg_place"]),
    }.get(sort, lambda x: x["avg_place"])
    heroes.sort(key=sort_key)

    return jsonify({"heroes": heroes, "total": len(heroes)})


@app.route("/upload", methods=["GET", "POST"])
def upload():
    if PUBLIC_MODE:
        return redirect(url_for("tier_list_page"))
    if request.method == "GET":
        return render_template("upload.html")

    file = request.files.get("excel")
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        return render_template("upload.html", error="請上傳 .xlsx 或 .xls 檔案")

    path = os.path.join(UPLOAD_DIR, secure_filename(file.filename))
    file.save(path)

    imported, skipped, failed, errors = _import_excel(path)
    msg = f"匯入 {imported} 筆，略過重複 {skipped} 筆"
    if failed:
        msg += f"，解析失敗 {failed} 筆"
    if errors:
        msg += f"（{errors[0]}）"
    return render_template("upload.html", success=msg)


def _build_card_lookups():
    """建立 BG 英雄 / 隨從名稱 → card_id 的反查字典。"""
    import re as _re

    bg_hero: dict = {}
    if os.path.exists(CARDS_CACHE):
        with open(CARDS_CACHE, encoding="utf-8") as f:
            cc = json.load(f)
        for cid, name in cc.items():
            if _re.match(r"(TB_BaconShop_HERO|BG\d+_HERO)", cid) and name not in bg_hero:
                bg_hero[name] = cid

    minion: dict = {}
    if os.path.exists(BG_MINIONS_CACHE):
        with open(BG_MINIONS_CACHE, encoding="utf-8") as f:
            for m in json.load(f):
                minion[m["name"]] = m["id"]

    return bg_hero, minion


_CARD_LOOKUPS: tuple | None = None   # lazy cache


def _get_card_lookups():
    global _CARD_LOOKUPS
    if _CARD_LOOKUPS is None:
        _CARD_LOOKUPS = _build_card_lookups()
    return _CARD_LOOKUPS


def _import_excel(path: str):
    """從 Excel 匯入紀錄到 SQLite（去重）。"""
    import re as _re
    from db import save_game, game_exists, start_time_exists
    from datetime import datetime

    bg_hero_map, minion_map = _get_card_lookups()

    def _resolve_hero_cid(name: str) -> str:
        return bg_hero_map.get(str(name or ""), "")

    def _resolve_board(board_text: str) -> list:
        result = []
        for raw in str(board_text or "").split("、"):
            raw = raw.strip()
            if not raw:
                continue
            is_golden = raw.startswith("★")
            clean = _re.sub(r"\[.*?\]|\(.*?\)", "", raw.lstrip("★")).strip()
            cid = minion_map.get(clean, "")
            result.append({"name": raw, "stats": raw, "card_id": cid,
                            "golden": is_golden})
        return result

    wb = openpyxl.load_workbook(path, read_only=True)
    imported = skipped = failed = 0
    errors = []
    seen_start_minutes: set[str] = set()   # 去重：每分鐘只保留第一筆（避免雙打隊友紀錄）

    def _parse_duration(s: str) -> int:
        if not s:
            return 0
        m = _re.search(r"(\d+)分(\d+)秒", str(s))
        if m:
            return int(m.group(1)) * 60 + int(m.group(2))
        return 0

    # 欄位順序（17 欄）：
    # 0日期 1版本 2模式 3英雄 4隊友英雄 5技能1 6技能2 7飾品1 8飾品2
    # 9名次 10板面 11回合 12金幣 13時長 14對手 15對手板面 16備註
    for sheet_name in ("單打紀錄", "雙打紀錄", "強力排組（單打）", "強力排組（雙打）"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # 標題列
                continue
            try:
                cols = list(row) + [None] * 20
                (dt_str, version, mode_label, hero_name, teammate_name,
                 p1, p2, t1, t2, placement, board_text,
                 turns, max_gold, duration_str, opponents,
                 opp_boards_text, note) = cols[:17]

                if not dt_str or not placement:
                    continue

                # openpyxl 日期欄位可能是 datetime 物件，也可能是字串
                if isinstance(dt_str, datetime):
                    start_time = dt_str
                else:
                    dt_s = str(dt_str).strip()
                    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
                        try:
                            start_time = datetime.strptime(dt_s, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        errors.append(f"日期解析失敗: {dt_str}")
                        failed += 1
                        continue

                game_id = f"excel_{start_time.strftime('%Y%m%d%H%M')}_{hero_name}"

                # 每分鐘只取第一筆（避免雙打隊友英雄重複記錄）
                ts_min = start_time.strftime('%Y%m%d%H%M')
                if ts_min in seen_start_minutes:
                    skipped += 1
                    continue
                seen_start_minutes.add(ts_min)

                if game_exists(game_id):
                    skipped += 1
                    continue

                # 對手英雄
                opp_heroes = [h.strip() for h in str(opponents or "").split("、") if h.strip()]

                # 對手板面
                opp_boards: dict = {}
                for line in str(opp_boards_text or "").split("\n"):
                    if "：" in line:
                        hname, mstr = line.split("：", 1)
                        opp_boards[hname.strip()] = _resolve_board(mstr)

                save_game(
                    game_id=game_id,
                    start_time=start_time,
                    end_time=None,
                    build_version=str(version or ""),
                    game_mode="duo" if "雙打" in str(mode_label or "") else "solo",
                    hero_card_id=_resolve_hero_cid(hero_name),
                    hero_name=str(hero_name or ""),
                    hero_power_ids=[],
                    hero_power_names=[p for p in [p1, p2] if p],
                    trinket_ids=[],
                    trinket_names=[t for t in [t1, t2] if t],
                    placement=int(placement),
                    final_board=_resolve_board(board_text),
                    penultimate_board=[],
                    turn_count=int(turns or 0),
                    max_gold=int(max_gold or 0),
                    duration_seconds=_parse_duration(duration_str),
                    opponent_heroes=opp_heroes,
                    opponent_boards=opp_boards,
                    teammate_hero_name=str(teammate_name or ""),
                )
                imported += 1
            except Exception as e:
                errors.append(f"Row {i}: {e}")
                failed += 1

    wb.close()
    return imported, skipped, failed, errors


# ── Hero Guide ────────────────────────────────────────────────────────────────
@app.route("/hero-guide")
def hero_guide_page():
    return render_template("hero_guide.html")


def _find_hero_power(cards, cid):
    """Try multiple hero power ID suffixes to find the correct power card."""
    # For t-suffix heroes (e.g. BG22_HERO_007t), also try stripping the t
    base_ids = [cid]
    if cid.endswith("t"):
        base_ids.append(cid[:-1])
    for base in base_ids:
        for sfx in ("p", "p2", "p3", "p4", "p5", "p_ALT"):
            candidate = base + sfx
            c = cards.get(candidate)
            if isinstance(c, dict) and c.get("type") == "HERO_POWER":
                return candidate, c
    return cid + "p", {}


@app.route("/api/hero-guide")
def api_hero_guide():
    # Prefer full cards file; fall back to lightweight pre-built heroes file
    if os.path.exists(HS_CARDS_FULL):
        cards = _load_json(HS_CARDS_FULL)
    elif os.path.exists(HS_BG_HEROES):
        cards = _load_json(HS_BG_HEROES)
    else:
        cards = {}
    q = (request.args.get("q") or "").strip().lower()
    heroes = []
    for cid, card in cards.items():
        if not isinstance(card, dict):
            continue
        if card.get("type") != "HERO" or not cid.startswith("BG") or "_SKIN" in cid:
            continue
        power_id, power = _find_hero_power(cards, cid)
        buddy_id = cid + "_Buddy"
        buddy = cards.get(buddy_id, {})
        entry = {
            "id": cid,
            "name": card.get("name", ""),
            "power_id": power_id,
            "power_name": power.get("name", ""),
            "power_cost": power.get("cost", 0),
            "power_text": _clean_card_text(power.get("text", "")),
            "buddy_id":     buddy_id if buddy else None,
            "buddy_name":   buddy.get("name", "") if buddy else None,
            "buddy_attack": buddy.get("attack") if buddy else None,
            "buddy_health": buddy.get("health") if buddy else None,
            "buddy_text":   _clean_card_text(buddy.get("text", "")) if buddy else None,
        }
        if q:
            searchable = (entry["name"] + entry["power_name"] + entry["power_text"] +
                          (entry["buddy_name"] or "") + (entry["buddy_text"] or "")).lower()
            if q not in searchable:
                continue
        heroes.append(entry)
    heroes.sort(key=lambda x: x["name"])
    return jsonify({"heroes": heroes, "total": len(heroes)})


# ── Spells Guide ───────────────────────────────────────────────────────────────
@app.route("/spells")
def spells_page():
    return render_template("spells.html")


if __name__ == "__main__":
    print("🌐 開啟瀏覽器：http://127.0.0.1:5000")
    app.run(debug=False, host="0.0.0.0", port=5000)
